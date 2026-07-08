"""Xuat bao cao Excel (.xlsx) co dinh dang: tieu de, header mau, ke bang,
so tien phan cach hang nghin, THU xanh / CHI do, dong tong cong.

Ho tro xuat song ngu Viet/Anh qua tham so lang="vi"|"en" (TEXT bang duoi).
CSV van giu lai qua tham so fmt=csv cho nhu cau xu ly may.
"""

from datetime import datetime, timedelta
from decimal import Decimal
from io import BytesIO
from pathlib import Path

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XlsxImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from PIL import Image as PilImage

# Kich thuoc thumbnail anh chung tu nhung vao Excel (pixel). Du lon de doc duoc
# so tien/ma GD tren anh chup man hinh ngan hang, khong qua lon lam nang file.
THUMB_MAX_WIDTH = 260
THUMB_MAX_HEIGHT = 340

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, size=14, color="1F4E78")
SUB_FONT = Font(size=10, color="595959", italic=True)
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
STRIPE_FILL = PatternFill("solid", fgColor="F2F7FB")
TOTAL_FILL = PatternFill("solid", fgColor="DDEBF7")
TOTAL_FONT = Font(bold=True, size=11)
THU_FONT = Font(bold=True, color="1E7E34")
CHI_FONT = Font(bold=True, color="C0392B")
MONEY_FORMAT = "#,##0"
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=False)
LEFT = Alignment(horizontal="left", vertical="center")
RIGHT = Alignment(horizontal="right", vertical="center")

# Tat ca chuoi hien thi song ngu tap trung o day - de them ngon ngu khac sau nay.
TEXT: dict[str, dict[str, str]] = {
    "vi": {
        "project_title": "BAO CAO TONG HOP DU AN",
        "project_sheet": "Tong hop du an",
        "transfer_title": "BAO CAO CHI TIET GIAO DICH THU/CHI",
        "transfer_sheet": "Chi tiet giao dich",
        "exported_at": "Xuat luc",
        "report_period": "Ky bao cao",
        "total": "TONG CONG",
        "total_thu": "TONG THU (da xac nhan)",
        "total_chi": "TONG CHI (da xac nhan)",
        "net": "CHENH LECH (THU - CHI)",
        "type_thu": "THU",
        "type_chi": "CHI",
        "type_unknown": "?",
        "no_image": "-",
    },
    "en": {
        "project_title": "PROJECT SUMMARY REPORT",
        "project_sheet": "Project Summary",
        "transfer_title": "TRANSACTION DETAIL REPORT (INCOME/EXPENSE)",
        "transfer_sheet": "Transaction Detail",
        "exported_at": "Exported at",
        "report_period": "Report period",
        "total": "TOTAL",
        "total_thu": "TOTAL INCOME (confirmed)",
        "total_chi": "TOTAL EXPENSE (confirmed)",
        "net": "NET (INCOME - EXPENSE)",
        "type_thu": "INCOME",
        "type_chi": "EXPENSE",
        "type_unknown": "?",
        "no_image": "-",
    },
}

PROJECT_HEADERS: dict[str, list[str]] = {
    "vi": [
        "Ma DA", "Ten du an", "Chu nhiem", "Phong ban",
        "Ngan sach (VND)", "Chi thuc te (VND)", "THU Telegram (VND)",
        "CHI Telegram (VND)", "Con lai (VND)", "% Su dung", "Trang thai",
    ],
    "en": [
        "Code", "Project Name", "Owner", "Department",
        "Budget (VND)", "Actual (VND)", "Telegram Income (VND)",
        "Telegram Expense (VND)", "Remaining (VND)", "% Used", "Status",
    ],
}

TRANSFER_HEADERS: dict[str, list[str]] = {
    "vi": [
        "Ngay nhan", "Thoi gian GD", "Du an", "Nguon", "Loai",
        "So tien (VND)", "Doi tac", "Ngan hang", "Ma GD / Noi dung CK",
        "Duyet", "Ghi chu", "Anh chung tu",
    ],
    "en": [
        "Received", "Transaction Time", "Project", "Source", "Type",
        "Amount (VND)", "Counterparty", "Bank", "Ref / Description",
        "Review", "Note", "Receipt Image",
    ],
}

REVIEW_LABELS: dict[str, dict[str, str]] = {
    "vi": {
        "pending_ai": "Cho AI",
        "pending_review": "Cho duyet",
        "confirmed": "Da xac nhan",
        "duplicate": "Nghi trung",
    },
    "en": {
        "pending_ai": "AI Processing",
        "pending_review": "Pending Review",
        "confirmed": "Confirmed",
        "duplicate": "Possible Duplicate",
    },
}

PERIOD_NAMES: dict[str, dict[str, str]] = {
    "vi": {
        "day": "Hom nay",
        "week": "Tuan nay",
        "month": "Thang nay",
        "year": "Nam nay",
        "custom": "Khoang tuy chon",
    },
    "en": {
        "day": "Today",
        "week": "This week",
        "month": "This month",
        "year": "This year",
        "custom": "Custom range",
    },
}


def _lang(lang: str) -> str:
    return lang if lang in TEXT else "vi"


def _init_sheet(
    ws: Worksheet,
    title: str,
    period_label: str,
    headers: list[str],
    widths: list[int],
    lang: str,
) -> int:
    """Ghi tieu de + ky bao cao + header; tra ve chi so dong du lieu dau tien."""
    last_col = get_column_letter(len(headers))

    ws.merge_cells(f"A1:{last_col}1")
    title_cell = ws["A1"]
    title_cell.value = title
    title_cell.font = TITLE_FONT
    title_cell.alignment = CENTER
    ws.row_dimensions[1].height = 26

    ws.merge_cells(f"A2:{last_col}2")
    sub_cell = ws["A2"]
    sub_cell.value = f"{period_label}  |  {TEXT[lang]['exported_at']} {datetime.now():%d/%m/%Y %H:%M}"
    sub_cell.font = SUB_FONT
    sub_cell.alignment = CENTER

    header_row = 3
    for index, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=index, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = BORDER
        cell.alignment = CENTER
        ws.column_dimensions[get_column_letter(index)].width = widths[index - 1]
    ws.row_dimensions[header_row].height = 20

    ws.freeze_panes = f"A{header_row + 1}"
    return header_row + 1


def _style_data_cell(cell, stripe: bool) -> None:
    cell.border = BORDER
    if stripe:
        cell.fill = STRIPE_FILL


def _receipt_thumbnail(file_path: str) -> tuple[BytesIO, int, int] | None:
    """Thu nho anh chung tu de nhung vao o Excel; hong/thieu file thi bo qua.

    Dung PNG (khong mat du lieu) thay vi JPEG nen manh - anh chup man hinh
    ngan hang la dang UI/chu nen JPEG de bi rang/mo chu so, PNG giu net hon
    va nen tot voi anh nhieu mang mau phang kieu nay.
    """
    try:
        path = Path(file_path)
        if not path.exists():
            return None
        with PilImage.open(path) as img:
            img = img.convert("RGB")
            img.thumbnail((THUMB_MAX_WIDTH, THUMB_MAX_HEIGHT), PilImage.LANCZOS)
            buffer = BytesIO()
            img.save(buffer, format="PNG", optimize=True)
            buffer.seek(0)
            return buffer, img.width, img.height
    except Exception:
        return None


def build_project_report_xlsx(rows: list[dict], period_label: str, lang: str = "vi") -> bytes:
    lang = _lang(lang)
    text = TEXT[lang]
    wb = Workbook()
    ws = wb.active
    ws.title = text["project_sheet"]

    headers = PROJECT_HEADERS[lang]
    widths = [10, 30, 18, 15, 17, 17, 17, 17, 17, 11, 12]
    row_index = _init_sheet(ws, text["project_title"], period_label, headers, widths, lang)

    money_columns = {5, 6, 7, 8, 9}
    for order, row in enumerate(rows):
        stripe = order % 2 == 1
        values = [
            row.get("code", ""),
            row.get("name", ""),
            row.get("owner", ""),
            row.get("department", ""),
            row.get("budget", 0),
            row.get("actual", 0),
            row.get("telegram_thu", 0),
            row.get("telegram_chi", 0),
            row.get("available", 0),
            row.get("usage_percent", 0),
            row.get("status", ""),
        ]
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=row_index, column=col, value=value)
            _style_data_cell(cell, stripe)
            if col in money_columns:
                cell.number_format = MONEY_FORMAT
                cell.alignment = RIGHT
            elif col == 10:
                cell.number_format = '0.0"%"'
                cell.alignment = RIGHT
            elif col in {1, 11}:
                cell.alignment = CENTER
            else:
                cell.alignment = LEFT
        row_index += 1

    if rows:
        ws.merge_cells(start_row=row_index, start_column=1, end_row=row_index, end_column=4)
        label_cell = ws.cell(row=row_index, column=1, value=text["total"])
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row_index, column=col)
            cell.fill = TOTAL_FILL
            cell.font = TOTAL_FONT
            cell.border = BORDER
        label_cell.alignment = CENTER
        for col, key in ((5, "budget"), (6, "actual"), (7, "telegram_thu"), (8, "telegram_chi"), (9, "available")):
            total_cell = ws.cell(row=row_index, column=col, value=sum(row.get(key, 0) or 0 for row in rows))
            total_cell.number_format = MONEY_FORMAT
            total_cell.alignment = RIGHT

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def build_transfer_report_xlsx(rows: list[dict], period_label: str, lang: str = "vi") -> bytes:
    lang = _lang(lang)
    text = TEXT[lang]
    review_labels = REVIEW_LABELS[lang]
    wb = Workbook()
    ws = wb.active
    ws.title = text["transfer_sheet"]

    headers = TRANSFER_HEADERS[lang]
    widths = [17, 17, 10, 11, 8, 15, 24, 18, 28, 13, 20, 39]
    image_column = len(headers)
    row_index = _init_sheet(ws, text["transfer_title"], period_label, headers, widths, lang)

    total_thu = Decimal("0")
    total_chi = Decimal("0")
    for order, row in enumerate(rows):
        stripe = order % 2 == 1
        received = row.get("received_at", "")
        try:
            received = f"{datetime.fromisoformat(received):%d/%m/%Y %H:%M}"
        except (TypeError, ValueError):
            received = str(received)

        transaction_type = row.get("transaction_type", "")
        type_label = {"thu": text["type_thu"], "chi": text["type_chi"]}.get(transaction_type, text["type_unknown"])
        amount = row.get("amount", "")
        confirmed = row.get("review_status") == "confirmed"
        if confirmed and isinstance(amount, (int, float)):
            if transaction_type == "thu":
                total_thu += Decimal(str(amount))
            elif transaction_type == "chi":
                total_chi += Decimal(str(amount))

        thumbnail = _receipt_thumbnail(row.get("file_path", "")) if row.get("file_path") else None
        values = [
            received,
            row.get("transacted_at", ""),
            row.get("project_code", "") or "-",
            row.get("source", ""),
            type_label,
            amount,
            row.get("counterparty", ""),
            row.get("bank_name", ""),
            row.get("reference", ""),
            review_labels.get(row.get("review_status", ""), row.get("review_status", "")),
            row.get("note", "") or row.get("caption", ""),
            "" if thumbnail else text["no_image"],
        ]
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=row_index, column=col, value=value)
            _style_data_cell(cell, stripe)
            if col == 5:
                cell.alignment = CENTER
                if transaction_type == "thu":
                    cell.font = THU_FONT
                elif transaction_type == "chi":
                    cell.font = CHI_FONT
            elif col == 6:
                cell.number_format = MONEY_FORMAT
                cell.alignment = RIGHT
            elif col in {1, 2, 3, 4, 10, 12}:
                cell.alignment = CENTER
            else:
                cell.alignment = LEFT

        if thumbnail:
            buffer, width, height = thumbnail
            image = XlsxImage(buffer)
            image.width = width
            image.height = height
            anchor = f"{get_column_letter(image_column)}{row_index}"
            ws.add_image(image, anchor)
            # Chieu cao dong theo point (1 px ~ 0.75 pt) de anh khong tran sang dong khac.
            ws.row_dimensions[row_index].height = height * 0.75 + 6
        row_index += 1

    summary = [
        (text["total_thu"], total_thu, THU_FONT),
        (text["total_chi"], total_chi, CHI_FONT),
        (text["net"], total_thu - total_chi, TOTAL_FONT),
    ]
    for label, value, font in summary:
        ws.merge_cells(start_row=row_index, start_column=1, end_row=row_index, end_column=5)
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row_index, column=col)
            cell.fill = TOTAL_FILL
            cell.border = BORDER
        label_cell = ws.cell(row=row_index, column=1, value=label)
        label_cell.font = font
        label_cell.alignment = RIGHT
        value_cell = ws.cell(row=row_index, column=6, value=float(value))
        value_cell.font = font
        value_cell.number_format = MONEY_FORMAT
        value_cell.alignment = RIGHT
        row_index += 1

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def period_display_label(period: str, start: datetime, end: datetime, lang: str = "vi") -> str:
    lang = _lang(lang)
    names = PERIOD_NAMES[lang]
    # end la can mo (exclusive) nen ngay cuoi cung trong ky la end - 1 ngay.
    display_end = end - timedelta(days=1)
    period_name = names.get(period, period)
    return f"{TEXT[lang]['report_period']}: {period_name} ({start:%d/%m/%Y} - {display_end:%d/%m/%Y})"


def project_scoped_label(
    project_name: str, project_code: str, start: datetime, end: datetime, lang: str = "vi"
) -> str:
    """Dong tieu de khi bao cao duoc loc theo 1 du an cu the - hien ten du an
    thay vi ten ky han chung chung, de nguoi mo file biet ngay day la bao cao
    cua du an nao."""
    lang = _lang(lang)
    display_end = end - timedelta(days=1)
    prefix = "Du an" if lang == "vi" else "Project"
    from_word = "Tu" if lang == "vi" else "From"
    return f"{prefix}: {project_name} ({project_code})  |  {from_word} {start:%d/%m/%Y} - {display_end:%d/%m/%Y}"
