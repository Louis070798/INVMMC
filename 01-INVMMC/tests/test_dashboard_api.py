from uuid import uuid4

from fastapi.testclient import TestClient

from invmmc.core.config import settings
from invmmc.main import app


def login_admin(client: TestClient) -> None:
    response = client.post(
        "/api/v1/auth/login",
        json={"email": settings.admin_email, "password": settings.admin_password},
    )
    assert response.status_code == 200


def test_dashboard_requires_login_then_loads_after_auth() -> None:
    with TestClient(app) as client:
        login_page = client.get("/dashboard")
        assert "loginForm" in login_page.text

        login_admin(client)
        page = client.get("/dashboard")
        summary = client.get("/api/v1/dashboard/summary?period=month")

    assert page.status_code == 200
    assert "INVMMC Finance Dashboard" in page.text
    assert summary.status_code == 200
    assert "kpis" in summary.json()


def test_project_crud_and_approval_flow() -> None:
    # Ma duy nhat moi lan chay de khong dung du an that / lan chay truoc bi crash.
    code = f"PRJT{uuid4().hex[:6].upper()}"
    project_id: str | None = None
    attachment_id: str | None = None
    with TestClient(app) as client:
        login_admin(client)
        try:
            created = client.post(
                "/api/v1/projects",
                json={
                    "code": code,
                    "name": "Du an test CRUD",
                    "owner": "Pytest",
                    "department": "QA",
                    "budget_amount": 5_000_000,
                },
            )
            assert created.status_code == 200
            body = created.json()
            project_id = body["id"]
            # Tao tu dashboard luon o trang thai cho duyet.
            assert body["status"] == "pending_approval"

            duplicate = client.post(
                "/api/v1/projects",
                json={
                    "code": code,
                    "name": "Trung ma",
                    "owner": "Pytest",
                    "department": "QA",
                    "budget_amount": 1_000_000,
                },
            )
            assert duplicate.status_code == 409

            edited = client.patch(
                f"/api/v1/projects/{project_id}",
                json={"name": "Du an test CRUD v2", "budget_amount": 7_000_000},
            )
            assert edited.status_code == 200
            assert edited.json()["name"] == "Du an test CRUD v2"
            assert edited.json()["status"] == "pending_approval"

            approved = client.patch(f"/api/v1/projects/{project_id}", json={"status": "active"})
            assert approved.status_code == 200
            assert approved.json()["status"] == "active"

            bad_status = client.patch(f"/api/v1/projects/{project_id}", json={"status": "bogus"})
            assert bad_status.status_code == 400

            # Du an da co chung tu thi khong xoa duoc.
            attachment = client.post(
                "/api/v1/attachments",
                data={"project_code": code, "transaction_type": "chi", "amount": "1000000"},
            )
            assert attachment.status_code == 200
            attachment_id = attachment.json()["id"]
            blocked = client.delete(f"/api/v1/projects/{project_id}")
            assert blocked.status_code == 409
            assert "project_has_data" in blocked.text

            removed = client.delete(f"/api/v1/attachments/{attachment_id}")
            assert removed.status_code == 200
            attachment_id = None

            deleted = client.delete(f"/api/v1/projects/{project_id}")
            assert deleted.status_code == 200
            listed = client.get("/api/v1/projects")
            assert all(row["id"] != project_id for row in listed.json())
            project_id = None
        finally:
            # Test chay tren DB that: don sach du lieu ke ca khi assert fail giua chung.
            if attachment_id:
                client.delete(f"/api/v1/attachments/{attachment_id}")
            if project_id:
                client.delete(f"/api/v1/projects/{project_id}")


XLSX_CONTENT_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def test_report_export_xlsx_default_and_csv_fallback() -> None:
    with TestClient(app) as client:
        locked = client.get("/api/v1/reports/export?period=week")
        assert locked.status_code == 401

        login_admin(client)
        xlsx = client.get("/api/v1/reports/export?period=week")
        csv_response = client.get("/api/v1/reports/export?period=week&fmt=csv")

    assert xlsx.status_code == 200
    assert xlsx.headers["content-type"].startswith(XLSX_CONTENT_TYPE)
    assert xlsx.content[:2] == b"PK"  # xlsx la file zip
    assert csv_response.status_code == 200
    assert csv_response.headers["content-type"].startswith("text/csv")
    assert "code,name,owner,department,budget,actual,telegram_thu,telegram_chi,available" in csv_response.text


def test_transfers_export_xlsx_default_and_csv_fallback() -> None:
    with TestClient(app) as client:
        locked = client.get("/api/v1/transfers/export?period=month")
        assert locked.status_code == 401

        login_admin(client)
        by_month = client.get("/api/v1/transfers/export?period=month")
        by_month_csv = client.get("/api/v1/transfers/export?period=month&fmt=csv")
        by_year = client.get("/api/v1/transfers/export?period=year")
        custom = client.get("/api/v1/transfers/export?period=custom&start_date=2026-01-01&end_date=2026-12-31")
        custom_missing = client.get("/api/v1/transfers/export?period=custom")
        custom_reversed = client.get(
            "/api/v1/transfers/export?period=custom&start_date=2026-12-31&end_date=2026-01-01"
        )

    assert by_month.status_code == 200
    assert by_month.headers["content-type"].startswith(XLSX_CONTENT_TYPE)
    assert by_month.content[:2] == b"PK"
    assert by_month_csv.status_code == 200
    assert by_month_csv.headers["content-type"].startswith("text/csv")
    assert "received_at,transacted_at,project_code,project_name" in by_month_csv.text
    assert by_year.status_code == 200
    assert custom.status_code == 200
    assert 'filename="invmmc-giao-dich-2026-01-01-to-2026-12-31.xlsx"' in custom.headers["content-disposition"]
    assert custom_missing.status_code == 422
    assert custom_reversed.status_code == 422


def test_export_lang_en_translates_headers_and_titles() -> None:
    from io import BytesIO

    from openpyxl import load_workbook

    with TestClient(app) as client:
        login_admin(client)
        project_en = client.get("/api/v1/reports/export?period=month&lang=en")
        project_vi = client.get("/api/v1/reports/export?period=month")
        transfer_en = client.get("/api/v1/transfers/export?period=month&lang=en")

    assert project_en.status_code == 200
    assert 'filename="invmmc-project-report-month.xlsx"' in project_en.headers["content-disposition"]
    wb_en = load_workbook(BytesIO(project_en.content))
    ws_en = wb_en.active
    assert ws_en["A1"].value == "PROJECT SUMMARY REPORT"
    assert ws_en["A3"].value == "Code"
    assert ws_en["B3"].value == "Project Name"

    # Mac dinh (khong truyen lang) van la tieng Viet - khong pha vo hanh vi cu.
    wb_vi = load_workbook(BytesIO(project_vi.content))
    ws_vi = wb_vi.active
    assert ws_vi["A1"].value == "BAO CAO TONG HOP DU AN"
    assert ws_vi["A3"].value == "Ma DA"

    assert transfer_en.status_code == 200
    wb_transfer_en = load_workbook(BytesIO(transfer_en.content))
    ws_transfer_en = wb_transfer_en.active
    assert ws_transfer_en["A1"].value == "TRANSACTION DETAIL REPORT (INCOME/EXPENSE)"
    assert ws_transfer_en["L3"].value == "Receipt Image"


def test_export_scoped_to_project_uses_project_name_and_dates_in_title_and_filename() -> None:
    from io import BytesIO

    from openpyxl import load_workbook

    code = f"PRJX{uuid4().hex[:6].upper()}"
    project_id: str | None = None
    with TestClient(app) as client:
        login_admin(client)
        try:
            created = client.post(
                "/api/v1/projects",
                json={
                    "code": code,
                    "name": "Chien dich test xuat theo du an",
                    "owner": "Pytest",
                    "department": "QA",
                    "budget_amount": 1_000_000,
                },
            )
            assert created.status_code == 200
            project_id = created.json()["id"]

            project_report = client.get(
                f"/api/v1/reports/export?period=custom&start_date=2026-01-01&end_date=2026-01-31&project_id={project_id}"
            )
            transfer_report = client.get(
                f"/api/v1/transfers/export?period=custom&start_date=2026-01-01&end_date=2026-01-31&project_id={project_id}"
            )

            assert project_report.status_code == 200
            expected_slug = f"invmmc-bao-cao-du-an-{code}-2026-01-01-to-2026-01-31.xlsx"
            assert f'filename="{expected_slug}"' in project_report.headers["content-disposition"]
            wb = load_workbook(BytesIO(project_report.content))
            ws = wb.active
            assert "Chien dich test xuat theo du an" in ws["A2"].value
            assert code in ws["A2"].value
            assert "01/01/2026" in ws["A2"].value and "31/01/2026" in ws["A2"].value

            assert transfer_report.status_code == 200
            expected_transfer_slug = f"invmmc-giao-dich-{code}-2026-01-01-to-2026-01-31.xlsx"
            assert f'filename="{expected_transfer_slug}"' in transfer_report.headers["content-disposition"]
            wb2 = load_workbook(BytesIO(transfer_report.content))
            ws2 = wb2.active
            assert "Chien dich test xuat theo du an" in ws2["A2"].value
        finally:
            if project_id:
                client.delete(f"/api/v1/projects/{project_id}")
